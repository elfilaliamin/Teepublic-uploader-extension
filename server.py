from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl

app = Flask(__name__)
CORS(app)

@app.route("/")
def home():
    return "Flask is running"

@app.route("/get-file", methods=["GET"])
def get_file():
    path = request.args.get("path")
    return send_file(path)

@app.route("/next-row", methods=["GET"])
def next_row():
    excel_path = request.args.get("path")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2):
        row_data = dict(zip(headers, [cell.value for cell in row]))

        if str(row_data.get("Status")).lower() == "done":
            continue

        return jsonify(row_data)

    return jsonify({"message": "No rows left"})


@app.route("/mark-done", methods=["POST"])
def mark_done():
    data = request.json
    excel_path = data.get("path")
    row_id = data.get("id")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    id_col = headers.index("Id") + 1
    status_col = headers.index("Status") + 1

    for r in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=r, column=id_col).value

        # If Id in Excel is number but JS sends string, compare as string
        if str(cell_value) == str(row_id):
            ws.cell(row=r, column=status_col).value = "Done"
            wb.save(excel_path)
            return jsonify({"status": "ok", "updated_id": row_id})

    return jsonify({"error": "Id not found"}), 404


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)
